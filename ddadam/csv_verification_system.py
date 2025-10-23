#!/usr/bin/env python3
"""
CSV自動訂正システム
JBAデータベースと照合してCSVファイルを自動訂正
"""

import streamlit as st
import pandas as pd
import requests
import json
from bs4 import BeautifulSoup
from datetime import datetime
import re
import unicodedata
from difflib import SequenceMatcher
import io
# import google.generativeai as genai  # AI機能は使用しない
import os
import concurrent.futures
import time
import threading

# ページ設定
st.set_page_config(
    page_title="CSV自動訂正システム",
    page_icon="🏀",
    layout="wide"
)

class JBAVerificationSystem:
    """JBA検証システム（requests + BeautifulSoupベース）"""
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36',
            'Accept': 'application/json',
            'Accept-Language': 'ja,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Origin': 'https://team-jba.jp',
            'Referer': 'https://team-jba.jp/organization/15250600/team/search',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'X-Requested-With': 'XMLHttpRequest'
        })
        self.logged_in = False
    
    def get_current_fiscal_year(self):
        """現在の年度を取得"""
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        if current_month >= 1:
            return str(current_year)
        else:
            return str(current_year - 1)
    
    def normalize_university_name(self, university_name):
        """大学名を正規化（柔軟な照合のため）"""
        if not university_name:
            return ""
        
        # 基本的な正規化
        normalized = university_name.strip()
        
        # よくある表記の統一
        replacements = {
            '白鷗大学': '白鴎大学',
            '白鴎大学': '白鴎大学',
            '白鷗': '白鴎',
            '白鴎': '白鴎',
            '大学': '大学',
            '学院': '学院',
            '短期大学': '短期大学',
            '短大': '短期大学'
        }
        
        for old, new in replacements.items():
            normalized = normalized.replace(old, new)
        
        return normalized
    
    def get_search_variations(self, university_name):
        """大学名の検索バリエーションを生成"""
        if not university_name:
            return []
        
        variations = [university_name.strip()]
        
        # 長い大学名の場合、短縮バリエーションも追加
        if len(university_name) > 6:  # 長い名前の場合
            # 語尾を段階的に削除
            suffixes_to_remove = ['体育会バスケットボール部', 'バスケットボール部', '体育会', '部']
            
            for suffix in suffixes_to_remove:
                if university_name.endswith(suffix):
                    base_name = university_name[:-len(suffix)].strip()
                    if base_name and len(base_name) > 2:  # 最低3文字以上
                        variations.append(base_name)
        
        # 重複を削除
        return list(set(variations))
    
    def login(self, email, password):
        """JBAサイトにログイン"""
        try:
            st.info("🔐 JBAサイトにログイン中...")
            
            login_page = self.session.get("https://team-jba.jp/login")
            soup = BeautifulSoup(login_page.content, 'html.parser')
            
            csrf_token = ""
            csrf_input = soup.find('input', {'name': '_token'})
            if csrf_input:
                csrf_token = csrf_input.get('value', '')
            
            login_data = {
                '_token': csrf_token,
                'login_id': email,
                'password': password
            }
            
            login_url = "https://team-jba.jp/login/done"
            login_response = self.session.post(login_url, data=login_data, allow_redirects=True)
            
            if "ログアウト" in login_response.text:
                st.success("✅ ログイン成功")
                self.logged_in = True
                return True
            else:
                st.error("❌ ログインに失敗しました")
                return False
                
        except Exception as e:
            st.error(f"❌ ログインエラー: {str(e)}")
            return False
    
    def search_teams_by_university(self, university_name):
        """大学名でチームを検索（柔軟な照合）"""
        try:
            if not self.logged_in:
                st.error("❌ ログインが必要です")
                return []
            
            current_year = self.get_current_fiscal_year()
            st.info(f"🔍 {university_name}の男子チームを検索中... ({current_year}年度)")
            
            # 大学名の正規化（柔軟な照合のため）
            normalized_university = self.normalize_university_name(university_name)
            st.info(f"🔍 正規化された大学名: {normalized_university}")
            
            # 正規化された大学名で検索
            search_university = normalized_university
            
            # 検索ページにアクセスしてCSRFトークンを取得
            search_url = "https://team-jba.jp/organization/15250600/team/search"
            search_page = self.session.get(search_url)
            
            if search_page.status_code != 200:
                st.error("❌ 検索ページにアクセスできません")
                return []
            
            soup = BeautifulSoup(search_page.content, 'html.parser')
            
            # CSRFトークンを取得
            csrf_token = ""
            csrf_input = soup.find('input', {'name': '_token'})
            if csrf_input:
                csrf_token = csrf_input.get('value', '')
            
            # JSON APIを使用した検索（男子チームのみ）
            search_data = {
                "limit": 100,
                "offset": 0,
                "searchLogic": "AND",
                "search": [
                    {"field": "fiscal_year", "type": "text", "operator": "is", "value": current_year},
                    {"field": "team_name", "type": "text", "operator": "contains", "value": search_university},
                    {"field": "competition_division_id", "type": "int", "operator": "is", "value": 1},
                    {"field": "team_search_out_of_range", "type": "int", "operator": "is", "value": 1}
                ]
            }
            
            form_data = {'request': json.dumps(search_data, ensure_ascii=False)}
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                'X-CSRF-Token': csrf_token,
                'X-Requested-With': 'XMLHttpRequest'
            }
            
            # 検索リクエストを送信（JSON APIとして）
            search_response = self.session.post(
                search_url, 
                data=form_data,
                headers=headers
            )
            
            if search_response.status_code != 200:
                st.error("❌ 検索リクエストが失敗しました")
                return []
            
            # JSONレスポンスを解析
            try:
                data = search_response.json()
                teams = []
                
                if data.get('status') == 'success' and 'records' in data:
                    for team_data in data['records']:
                        # 男子チームのみを対象
                        if team_data.get('team_gender_id') == '男子':
                            teams.append({
                                'id': team_data.get('id', ''),
                                'name': team_data.get('team_name', ''),
                                'url': f"https://team-jba.jp/organization/15250600/team/{team_data.get('id', '')}/detail"
                            })
                
                st.success(f"✅ {university_name}の男子チーム: {len(teams)}件見つかりました")
                return teams
                
            except Exception as e:
                st.error(f"❌ 検索結果の解析に失敗しました: {str(e)}")
                return []
            
        except Exception as e:
            st.error(f"❌ チーム検索エラー: {str(e)}")
            return []
    
    def _search_teams_by_university_silent(self, university_name):
        """大学名でチームを検索（静かな実行版 - st.*出力なし）"""
        try:
            if not self.logged_in:
                return []
            
            current_year = self.get_current_fiscal_year()
            
            # 大学名の正規化（柔軟な照合のため）
            normalized_university = self.normalize_university_name(university_name)
            
            # 正規化された大学名で検索
            search_university = normalized_university
            
            # 検索ページにアクセスしてCSRFトークンを取得
            search_url = "https://team-jba.jp/organization/15250600/team/search"
            search_page = self.session.get(search_url)
            
            if search_page.status_code != 200:
                return []
            
            soup = BeautifulSoup(search_page.content, 'html.parser')
            
            # CSRFトークンを取得
            csrf_token = ""
            csrf_input = soup.find('input', {'name': '_token'})
            if csrf_input:
                csrf_token = csrf_input.get('value', '')
            
            # JSON APIを使用した検索（男子チームのみ）
            search_data = {
                "limit": 100,
                "offset": 0,
                "searchLogic": "AND",
                "search": [
                    {"field": "fiscal_year", "type": "text", "operator": "is", "value": current_year},
                    {"field": "team_name", "type": "text", "operator": "contains", "value": search_university},
                    {"field": "competition_division_id", "type": "int", "operator": "is", "value": 1},
                    {"field": "team_search_out_of_range", "type": "int", "operator": "is", "value": 1}
                ]
            }
            
            form_data = {'request': json.dumps(search_data, ensure_ascii=False)}
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                'X-CSRF-Token': csrf_token,
                'X-Requested-With': 'XMLHttpRequest'
            }
            
            # 検索リクエストを送信（JSON APIとして）
            search_response = self.session.post(
                search_url, 
                data=form_data,
                headers=headers
            )
            
            if search_response.status_code != 200:
                return []
            
            # JSONレスポンスを解析
            try:
                data = search_response.json()
                teams = []
                
                if data.get('status') == 'success' and 'records' in data:
                    for team_data in data['records']:
                        # 男子チームのみを対象
                        if team_data.get('team_gender_id') == '男子':
                            teams.append({
                                'id': team_data.get('id', ''),
                                'name': team_data.get('team_name', ''),
                                'url': f"https://team-jba.jp/organization/15250600/team/{team_data.get('id', '')}/detail"
                            })
                
                return teams
                
            except Exception as e:
                return []
            
        except Exception as e:
            return []

    def get_team_members(self, team_url):
        """チームのメンバー情報を取得（男子チームのみ）"""
        try:
            st.info(f"📊 チームメンバー情報を取得中...")
            st.write(f"🔍 チームURL: {team_url}")
            
            # チーム詳細ページにアクセス
            team_page = self.session.get(team_url)
            
            if team_page.status_code != 200:
                st.error(f"❌ チームページにアクセスできません (Status: {team_page.status_code})")
                return {"team_name": "Error", "members": []}
            
            soup = BeautifulSoup(team_page.content, 'html.parser')
            
            # チーム名を取得
            team_name = "Unknown Team"
            title_element = soup.find('title')
            if title_element:
                team_name = title_element.get_text(strip=True)
            
            st.write(f"🔍 チーム名: {team_name}")

            # メンバー情報を抽出（男子チームのメンバーテーブルを特定）
            members = []
            
            tables = soup.find_all('table')

            # 男子チームのメンバーテーブルを探す（3列のテーブルを探す）
            member_table = None
            for i, table in enumerate(tables):
                rows = table.find_all('tr')
                if len(rows) > 10:  # メンバーテーブルは通常10行以上
                    # 最初の行に「メンバーID / 氏名 / 生年月日」があるかチェック
                    first_row_cells = rows[0].find_all(['td', 'th'])
                    if len(first_row_cells) >= 3:
                        first_cell = first_row_cells[0].get_text(strip=True)
                        second_cell = first_row_cells[1].get_text(strip=True)
                        third_cell = first_row_cells[2].get_text(strip=True)
                        if "メンバーID" in first_cell and "氏名" in second_cell and "生年月日" in third_cell:
                            member_table = table
                            break

            if member_table:
                rows = member_table.find_all('tr')
                for row in rows[1:]:  # ヘッダー行をスキップ
                    cells = row.find_all(['td', 'th'])
                    if len(cells) >= 3:
                        member_id = cells[0].get_text(strip=True)
                        name = cells[1].get_text(strip=True)
                        birth_date = cells[2].get_text(strip=True)
                        
                        # メンバーIDが数字で、名前が空でない場合のみ追加
                        if member_id.isdigit() and name and name != "氏名":
                            # 選手詳細ページのリンクを取得
                            detail_link = None
                            name_cell = cells[1]
                            link = name_cell.find('a')
                            if link and link.get('href'):
                                detail_link = link.get('href')
                                # 相対URLを絶対URLに変換
                                if detail_link.startswith('/'):
                                    detail_link = f"https://team-jba.jp{detail_link}"
                            
                            members.append({
                                "member_id": member_id,
                                "name": name,
                                "birth_date": birth_date,
                                "detail_url": detail_link
                            })

            return {
                "team_name": team_name,
                "members": members
            }
            
        except Exception as e:
            st.error(f"❌ メンバー取得エラー: {str(e)}")
            import traceback
            st.write(f"**エラー詳細**: {traceback.format_exc()}")
            return {"team_name": "Error", "team_url": team_url, "members": []}
    
    def _get_team_members_silent(self, team_url):
        """チームのメンバー情報を取得（静かな実行版 - st.*出力なし）"""
        try:
            # チーム詳細ページにアクセス
            team_page = self.session.get(team_url)
            
            if team_page.status_code != 200:
                return {"team_name": "Error", "members": []}
            
            soup = BeautifulSoup(team_page.content, 'html.parser')
            
            # チーム名を取得
            team_name = "Unknown Team"
            title_element = soup.find('title')
            if title_element:
                team_name = title_element.get_text(strip=True)
            
            # メンバー情報を取得
            members = []
            
            # 選手一覧のテーブルを探す
            member_tables = soup.find_all('table', class_='table')
            
            for table in member_tables:
                rows = table.find_all('tr')
                
                for row in rows[1:]:  # ヘッダー行をスキップ
                    cells = row.find_all(['td', 'th'])
                    
                    if len(cells) >= 3:  # 最低限の情報がある行のみ処理
                        # 選手名のリンクを探す
                        name_link = row.find('a', href=re.compile(r'/player/\d+'))
                        
                        if name_link:
                            player_name = name_link.get_text(strip=True)
                            detail_url = name_link['href']
                            
                            if not detail_url.startswith('http'):
                                detail_url = f"https://team-jba.jp{detail_url}"
                            
                            # その他の情報を取得
                            position = ""
                            grade = ""
                            height = ""
                            weight = ""
                            
                            for i, cell in enumerate(cells):
                                cell_text = cell.get_text(strip=True)
                                
                                # ポジション（通常は2番目のカラム）
                                if i == 1 and cell_text and cell_text not in ['選手名', '氏名']:
                                    position = cell_text
                                
                                # 学年（通常は3番目のカラム）
                                elif i == 2 and cell_text and cell_text not in ['学年', '年']:
                                    grade = cell_text
                                
                                # 身長・体重の情報を探す
                                if 'cm' in cell_text:
                                    height = cell_text
                                elif 'kg' in cell_text:
                                    weight = cell_text
                            
                            members.append({
                                "name": player_name,
                                "position": position,
                                "grade": grade,
                                "height": height,
                                "weight": weight,
                                "detail_url": detail_url
                            })
            
            return {
                "team_name": team_name,
                "members": members
            }
            
        except Exception as e:
            return {"team_name": "Error", "members": []}
    
    def get_player_details(self, detail_url):
        """選手詳細ページから身長・体重などの詳細情報を取得"""
        try:
            if not detail_url:
                return {}
            
            st.info(f"🔍 選手詳細情報を取得中: {detail_url}")
            
            # 選手詳細ページにアクセス
            detail_page = self.session.get(detail_url)
            
            if detail_page.status_code != 200:
                st.warning(f"⚠️ 選手詳細ページにアクセスできません (Status: {detail_page.status_code})")
                return {}
            
            soup = BeautifulSoup(detail_page.content, 'html.parser')
            
            # 選手詳細情報を抽出
            player_details = {}
            
            # 身長・体重情報を探す
            # 一般的なパターンを試す
            height_patterns = [
                r'身長[：:]\s*(\d+\.?\d*)\s*cm',
                r'身長[：:]\s*(\d+\.?\d*)\s*センチ',
                r'Height[：:]\s*(\d+\.?\d*)\s*cm'
            ]
            
            weight_patterns = [
                r'体重[：:]\s*(\d+\.?\d*)\s*kg',
                r'体重[：:]\s*(\d+\.?\d*)\s*キロ',
                r'Weight[：:]\s*(\d+\.?\d*)\s*kg'
            ]
            
            # テーブルから情報を抽出
            tables = soup.find_all('table')
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) >= 2:
                        label = cells[0].get_text(strip=True)
                        value = cells[1].get_text(strip=True)
                        
                        # 身長情報（JBAの「身長（競技者用）」に対応）
                        if '身長' in label or 'Height' in label:
                            # 数値部分を抽出
                            import re
                            height_match = re.search(r'(\d+\.?\d*)', value)
                            if height_match and value.strip():  # 空でない場合のみ
                                player_details['height'] = height_match.group(1)
                        
                        # 体重情報（JBAの「体重（競技者用）」に対応）
                        elif '体重' in label or 'Weight' in label:
                            # 数値部分を抽出
                            import re
                            weight_match = re.search(r'(\d+\.?\d*)', value)
                            if weight_match and value.strip():  # 空でない場合のみ
                                player_details['weight'] = weight_match.group(1)
                        
                        # ポジション情報
                        elif 'ポジション' in label or 'Position' in label:
                            player_details['position'] = value
                        
                        # 出身校情報
                        elif '出身校' in label or '出身' in label:
                            player_details['school'] = value
                        
                        # 学年情報
                        elif '学年' in label or 'Grade' in label:
                            player_details['grade'] = value
                        
                        # ユニフォーム番号
                        elif 'ユニフォーム番号' in label or '背番号' in label:
                            player_details['uniform_number'] = value
            
            # テーブルで見つからない場合は、ページ全体から正規表現で検索
            if 'height' not in player_details or 'weight' not in player_details:
                page_text = soup.get_text()
                
                # 身長を検索
                for pattern in height_patterns:
                    import re
                    match = re.search(pattern, page_text)
                    if match:
                        player_details['height'] = match.group(1)
                        break
                
                # 体重を検索
                for pattern in weight_patterns:
                    import re
                    match = re.search(pattern, page_text)
                    if match:
                        player_details['weight'] = match.group(1)
                        break
            
            return player_details
            
        except Exception as e:
            st.warning(f"⚠️ 選手詳細取得エラー: {str(e)}")
            return {}
    

    def normalize_name(self, name):
        """名前の正規化"""
        if not name or pd.isna(name):
            return ""
        
        name = str(name)
        
        # 1. 全角・半角統一
        name = unicodedata.normalize('NFKC', name)
        
        # 2. 記号・スペースの正規化（全角スペースも含む）
        name = re.sub(r'[・･、，,\.\s　]+', '', name)
        
        # 3. 大文字小文字統一
        name = name.lower()
        
        # 4. よくある表記揺れの統一
        name = re.sub(r'[ー−‐—–]', '', name)  # 長音符、ハイフン、エムダッシュ、エンダッシュ除去
        
        return name

    def calculate_similarity(self, name1, name2):
        """名前の類似度を計算"""
        if not name1 or not name2:
            return 0.0
        
        # 正規化
        norm_name1 = self.normalize_name(name1)
        norm_name2 = self.normalize_name(name2)
        
        if norm_name1 == norm_name2:
            return 1.0
        
        # 基本的な類似度
        basic_similarity = SequenceMatcher(None, norm_name1, norm_name2).ratio()
        
        return basic_similarity
    
    def show_name_differences(self, name1, name2):
        """名前の微妙な違いを視覚的に表示"""
        if not name1 or not name2:
            return ""
        
        # 正規化
        norm_name1 = self.normalize_name(name1)
        norm_name2 = self.normalize_name(name2)
        
        if norm_name1 == norm_name2:
            return "✅ 完全一致"
        
        # 文字単位での差分を表示
        matcher = SequenceMatcher(None, norm_name1, norm_name2)
        differences = []
        
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                differences.append(norm_name1[i1:i2])
            elif tag == 'delete':
                differences.append(f"❌{norm_name1[i1:i2]}❌")
            elif tag == 'insert':
                differences.append(f"➕{norm_name2[j1:j2]}➕")
            elif tag == 'replace':
                differences.append(f"🔄{norm_name1[i1:i2]}→{norm_name2[j1:j2]}🔄")
        
        result = "".join(differences)
        return f"🔍 差分: {result}"

    def verify_player_info(self, player_name, birth_date, university, get_details=False, threshold=1.0):
        """個別選手情報の照合（男子チームのみ）"""
        try:
            st.write(f"🔍 選手照合: {player_name}, 大学: {university}")
            
            # 大学名の検索バリエーションを生成
            search_variations = self.get_search_variations(university)
            st.write(f"🔍 検索バリエーション: {search_variations}")
            
            all_matched_members = []  # すべてのマッチ候補を保存
            
            teams = []
            for variation in search_variations:
                st.write(f"🔍 チーム検索開始: {variation}")
                teams = self.search_teams_by_university(variation)
                st.write(f"🔍 検索結果: {len(teams)}チーム見つかりました")
                
                if teams:
                    st.success(f"✅ {variation}でチームが見つかりました")
                    break
                else:
                    st.info(f"❌ {variation}ではチームが見つかりませんでした")
            
            if not teams:
                st.warning(f"❌ {university}の男子チームが見つかりませんでした")
                return {"status": "not_found", "message": f"{university}の男子チームが見つかりませんでした"}

            # 各チームのメンバー情報を取得して照合
            for team in teams:
                st.write(f"🔍 チーム: {team['name']} のメンバーを取得中...")
                team_data = self.get_team_members(team['url'])
                
                if team_data and team_data["members"]:
                    st.write(f"🔍 メンバー数: {len(team_data['members'])}人")
                    
                    for i, member in enumerate(team_data["members"]):
                        st.write(f"  - メンバー{i+1}: {member['name']}")
                        
                        # 名前の類似度チェック
                        name_similarity = self.calculate_similarity(player_name, member["name"])

                        # デバッグ情報を表示
                        st.write(f"  - JBA選手: {member['name']}")
                        st.write(f"  - 名前類似度: {name_similarity:.3f}")
                        
                        # 微妙な違いを表示（0.6以上の候補のみ）
                        if name_similarity >= 0.6:
                            diff_info = self.show_name_differences(player_name, member["name"])
                            st.write(f"  - {diff_info}")

                        # 第1段階: 0.6の閾値で候補を探す
                        if name_similarity >= 0.6:
                            st.info(f"🔍 候補発見: {member['name']} (類似度: {name_similarity:.3f})")
                            
                            # 第2段階: 1.0の閾値で完全一致を確認
                            if name_similarity >= 1.0:
                                st.success(f"✅ 完全一致: {member['name']}")
                                
                                # 詳細情報を取得する場合
                                if get_details and member.get("detail_url"):
                                    player_details = self.get_player_details(member["detail_url"])
                                    member.update(player_details)
                                
                                return {
                                    "status": "match",
                                    "jba_data": member,
                                    "similarity": name_similarity
                                }
                            
                            # 0.6以上1.0未満の候補も保存（最終的に返す可能性）
                            elif name_similarity >= 0.6 and name_similarity < 1.0:
                                st.info(f"📝 候補保存: {member['name']} (類似度: {name_similarity:.3f})")
                                
                                if get_details and member.get("detail_url"):
                                    player_details = self.get_player_details(member["detail_url"])
                                    member.update(player_details)
                                
                                all_matched_members.append({
                                    "status": "partial_match",
                                    "jba_data": member,
                                    "similarity": name_similarity,
                                    "message": f"部分一致: {member['name']} (類似度: {name_similarity:.3f})"
                                })
                else:
                    st.warning(f"❌ チーム {team['name']} のメンバー情報が取得できませんでした")

            # 完全一致を優先し、なければ部分一致を返す
            if all_matched_members:
                # 完全一致（類似度1.0）を優先
                exact_matches = [m for m in all_matched_members if m["similarity"] >= 1.0]
                if exact_matches:
                    st.info(f"🎯 完全一致候補: {len(exact_matches)}件")
                    return exact_matches[0]  # 最初の完全一致を返す
                
                # 部分一致（類似度0.6以上1.0未満）を返す
                partial_matches = [m for m in all_matched_members if m["similarity"] >= 0.6 and m["similarity"] < 1.0]
                if partial_matches:
                    st.info(f"📝 部分一致候補: {len(partial_matches)}件")
                    return partial_matches[0]  # 最初の部分一致を返す
                
                # その他の候補
                st.info(f"🔍 その他候補: {len(all_matched_members)}件")
                return all_matched_members[0]

            return {"status": "not_found", "message": "JBAデータベースに該当する選手が見つかりませんでした"}

        except Exception as e:
            return {"status": "error", "message": f"照合エラー: {str(e)}"}

# AI機能は使用しないため削除
    
# AI機能は使用しないため削除
    
# AI機能は使用しないため削除

class DataValidator:
    """データ検証システム（AI機能なし）"""
    
    def __init__(self, gemini_api_key=None):
        # AI機能は使用しない
        pass
    
    def validate_weight(self, weight):
        """体重の妥当性を評価（AI機能なし）"""
        if not weight:
            return True, []
        
        # シンプルな範囲チェック
        try:
            weight_value = float(weight)
            if 45 <= weight_value <= 140:
                return True, []
            else:
                return False, [f"体重が範囲外です: {weight}kg (45-140kgの範囲で入力してください)"]
        except (ValueError, TypeError):
            return False, [f"体重が数値ではありません: {weight}"]
    
    def validate_and_correct_school(self, school_name):
        """出身校の妥当性を評価（AI機能なし）"""
        if not school_name or school_name.strip() == "":
            return True, [], None
        
        # シンプルな文字列チェック
        school_name = str(school_name).strip()
        if len(school_name) < 2:
            return False, ["学校名が短すぎます"], None
        
        return True, [], None
    
    def validate_uniform_number(self, uniform_number):
        """背番号の妥当性を評価（AI機能なし）"""
        if not uniform_number:
            return True, []
        
        # 背番号は数字のみのシンプル検証
        try:
            num = int(uniform_number)
            if 1 <= num <= 99:
                return True, []
            else:
                return False, ["背番号は1〜99の範囲である必要があります"]
        except ValueError:
            return False, ["背番号は数字である必要があります"]
    
    def validate_player_data(self, player_data):
        """体重・出身校・背番号の検証（AI機能なし）"""
        all_issues = []
        
        # 体重の検証
        weight = player_data.get('weight')
        if weight:
            is_valid_weight, weight_issues = self.validate_weight(weight)
            all_issues.extend(weight_issues)
        
        # 出身校の検証
        school = player_data.get('school')
        if school:
            is_valid_school, school_issues, _ = self.validate_and_correct_school(school)
            all_issues.extend(school_issues)
        
        # 背番号の検証
        uniform_number = player_data.get('uniform_number')
        if uniform_number:
            is_valid_uniform, uniform_issues = self.validate_uniform_number(uniform_number)
            all_issues.extend(uniform_issues)
        
        return len(all_issues) == 0, all_issues

class FastCSVCorrectionSystem:
    """CSV訂正システム（改版）"""
    
    def __init__(self, jba_system, gemini_api_key=None, max_workers=5):
        self.jba_system = jba_system
        self.validator = DataValidator(gemini_api_key)
        self.max_workers = max_workers
        self.lock = threading.Lock()
    
    def _preload_university_data(self, university_name):
        """大学のチーム情報を事前に全て取得（1回だけ実行）"""
        if university_name in self.university_teams_data:
            return self.university_teams_data[university_name]
        
        # チーム検索（検索バリエーション対応）- 静かな実行
        search_variations = self.jba_system.get_search_variations(university_name)
        teams = []
        
        for variation in search_variations:
            teams = self.jba_system._search_teams_by_university_silent(variation)
            if teams:
                break
        
        if not teams:
            with self.lock:
                self.university_teams_data[university_name] = None
            return None
        
        # 各チームのメンバーを取得 - 静かな実行
        teams_data = {}
        
        for team in teams:
            team_id = team['id']
            team_url = team['url']
            
            # 既にキャッシュにあれば使用
            if team_url in self.team_members_cache:
                team_data = self.team_members_cache[team_url]
            else:
                team_data = self.jba_system._get_team_members_silent(team_url)
                with self.lock:
                    self.team_members_cache[team_url] = team_data
            
            teams_data[team_id] = {
                'team_name': team['name'],
                'team_url': team_url,
                'members': team_data.get('members', [])
            }
        
        with self.lock:
            self.university_teams_data[university_name] = teams_data
        
        return teams_data
    
    def _process_single_player(self, row_data):
        """単一選手を処理（訂正必要な場合のみ情報を詰める）"""
        index, row, university_name, threshold = row_data
        
        try:
            player_name = None
            name_column = None
            name_columns = ['選手名', '氏名', 'name', 'Name']
            
            for col in name_columns:
                if col in row.index and pd.notna(row[col]):
                    player_name = str(row[col]).strip()
                    name_column = col
                    break
            
            if not player_name:
                return {
                    'index': index,
                    'original_data': row.to_dict(),
                    'status': 'missing_data',
                    'corrections': {},
                    'jba_data': {},
                    'validation_warnings': [],
                    'has_correction': False
                }
            
            # JBAデータベースと照合
            verification_result = self.jba_system.verify_player_info(
                player_name, None, university_name, get_details=True, threshold=threshold
            )
            
            result = {
                'index': index,
                'original_data': row.to_dict(),
                'verification_result': verification_result,
                'status': verification_result.get('status', 'error'),
                'corrections': {},
                'jba_data': {},
                'validation_warnings': [],
                'has_correction': False
            }
            
            # jba_data を事前に初期化
            jba_data = {}
            
            if verification_result.get('status') in ['match', 'partial_match']:
                jba_data = verification_result.get('jba_data', {})
                result['jba_data'] = jba_data
                
                # 名前が異なる場合のみ訂正
                if jba_data.get('name') and jba_data['name'] != player_name:
                    result['corrections'][name_column] = jba_data['name']
                    result['has_correction'] = True
                
                # 体重：JBAにあれば優先し、元データと異なる場合のみ訂正
                if jba_data.get('weight') and str(jba_data['weight']).strip():
                    weight_value = str(jba_data['weight']).strip()
                    weight_match = re.search(r'(\d+\.?\d*)', weight_value)
                    if weight_match:
                        extracted_weight = weight_match.group(1)
                        try:
                            original_weight = float(row.get('体重', 0))
                            jba_weight = float(extracted_weight)
                            if original_weight != jba_weight:
                                result['corrections']['体重'] = extracted_weight
                                result['has_correction'] = True
                        except (ValueError, TypeError):
                            pass
                
                # 学年：JBAに記載があれば、数字だけを抽出し、元データと異なる場合のみ訂正
                if jba_data.get('grade') and str(jba_data['grade']).strip():
                    grade_value = str(jba_data['grade']).strip()
                    grade_match = re.search(r'(\d+)', grade_value)
                    if grade_match:
                        extracted_grade = grade_match.group(1)
                        try:
                            original_grade = str(row.get('学年', '')).strip()
                            if original_grade.isdigit():
                                original_grade_num = original_grade
                            else:
                                grade_num_match = re.search(r'(\d+)', original_grade)
                                original_grade_num = grade_num_match.group(1) if grade_num_match else original_grade
                            
                            if original_grade_num != extracted_grade:
                                result['corrections']['学年'] = extracted_grade
                                result['has_correction'] = True
                        except:
                            pass
                
                # 身長：JBAに記載があれば、数字だけを抽出し、元データと異なる場合のみ訂正
                if jba_data.get('height') and str(jba_data['height']).strip():
                    height_value = str(jba_data['height']).strip()
                    height_match = re.search(r'(\d+\.?\d*)', height_value)
                    if height_match:
                        extracted_height = height_match.group(1)
                        try:
                            original_height = float(row.get('身長', 0))
                            jba_height = float(extracted_height)
                            if original_height != jba_height:
                                result['corrections']['身長'] = extracted_height
                                result['has_correction'] = True
                        except (ValueError, TypeError):
                            pass
                
                # 元データの異常値をAIで検出（JBAにデータがない場合のみ）
                if not jba_data.get('weight') and not jba_data.get('height'):
                    validation_warnings = []  # AI機能は使用しない
                    result['validation_warnings'] = validation_warnings
            else:
                # JBA登録なし・未発見の場合も警告をチェック
                validation_warnings = []  # AI機能は使用しない
                result['validation_warnings'] = validation_warnings
            
            return result
        
        except Exception as e:
            import traceback
            return {
                'index': index,
                'original_data': row.to_dict(),
                'status': 'error',
                'corrections': {},
                'jba_data': {},
                'validation_warnings': [f'エラー: {str(e)}'],
                'has_correction': False
            }
    
    def _validate_player_data_with_ai(self, row, jba_data):
        """元データの異常値を検出（JBAに記載がない場合のみ）"""
        warnings = []
        
        # 体重：JBAに記載がない場合のみ許容範囲でチェック
        if not jba_data.get('weight') and pd.notna(row.get('体重')):
            weight = row.get('体重')
            try:
                weight_value = float(weight)
                if weight_value < 45 or weight_value > 140:
                    warnings.append(f"⚠️ 体重が許容範囲外: {weight_value}kg (許容範囲: 45-140kg)")
            except (ValueError, TypeError):
                warnings.append(f"⚠️ 体重が数値ではない: {weight}")
        
        return warnings
    
    def process_csv_file_parallel(self, df, university_name, threshold=1.0):
        """CSVファイルを並列処理で高速に処理"""
        
        st.info("ステップ1: データを並列処理中...")
        
        process_data = [
            (index, row, university_name, threshold)
            for index, row in df.iterrows()
        ]
        
        results = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        start_time = time.time()
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(self._process_single_player, data): data[0] for data in process_data}
            
            completed = 0
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                results.append(result)
                
                completed += 1
                progress = completed / len(futures)
                progress_bar.progress(progress)
                
                player_name = result['original_data'].get('選手名', 'Unknown')
                status_text.text(f"処理中: {completed}/{len(futures)} - {player_name}")
        
        elapsed_time = time.time() - start_time
        
        progress_bar.progress(1.0)
        status_text.text("✅ 処理完了")
        
        results.sort(key=lambda x: x['index'])
        
        # ★ 結果サマリーを表示
        st.success(f"✅ {len(df)}行を{elapsed_time:.2f}秒で処理しました")
        
        # 統計情報
        matched = sum(1 for r in results if r['status'] == 'match')
        partial = sum(1 for r in results if r['status'] == 'partial_match')
        warnings_count = sum(len(r.get('validation_warnings', [])) for r in results)
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("JBA一致", matched)
        with col2:
            st.metric("部分一致", partial)
        with col3:
            st.metric("⚠️ 警告", warnings_count)
        with col4:
            st.metric("処理時間", f"{elapsed_time:.2f}秒")
        
        return results
    
    def create_corrected_csv(self, df, results):
        """修正版CSVを作成（元の列順を保持、セル形式を保持）"""
        corrected_df = df.copy()
        
        for result in results:
            # 訂正がある場合のみ処理
            if not result.get('has_correction'):
                continue
            
            index = result['index']
            corrections = result.get('corrections', {})
            
            if not corrections:
                continue
            
            # 各修正項目をCSVに反映（列順は変わらない）
            for csv_col, corrected_value in corrections.items():
                if csv_col not in corrected_df.columns:
                    # 列が存在しない場合はスキップ（追加しない）
                    continue
                
                # 修正値を適用
                corrected_df.at[index, csv_col] = corrected_value
        
        return corrected_df
    
    def create_colored_excel(self, df, results):
        """色付きExcelファイルを作成（修正箇所は赤、警告は黄色、全体中央揃え）"""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='修正済み')
            
            ws = writer.sheets['修正済み']
            
            # スタイル定義
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            white_font = Font(color='FFFFFF', bold=True)
            black_font = Font(color='000000')
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            columns = list(df.columns)
            
            # 全セルを中央揃えに
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
                for cell in row:
                    cell.alignment = center_alignment
            
            for result in results:
                row_index = result['index'] + 2
                
                # 訂正がある場合：赤色
                if result.get('has_correction'):
                    for col_name, corrected_value in result['corrections'].items():
                        if col_name in columns:
                            col_index = columns.index(col_name) + 1
                            cell = ws.cell(row=row_index, column=col_index)
                            cell.fill = red_fill
                            cell.font = white_font
                            cell.alignment = center_alignment
                
                # 警告がある場合：黄色
                if result.get('validation_warnings'):
                    for warning in result['validation_warnings']:
                        if '体重' in warning and '体重' in columns:
                            col_index = columns.index('体重') + 1
                            cell = ws.cell(row=row_index, column=col_index)
                            cell.fill = yellow_fill
                            cell.font = black_font
                            cell.alignment = center_alignment
                        elif '身長' in warning and '身長' in columns:
                            col_index = columns.index('身長') + 1
                            cell = ws.cell(row=row_index, column=col_index)
                            cell.fill = yellow_fill
                            cell.font = black_font
                            cell.alignment = center_alignment
                        elif '出身校' in warning and '出身校' in columns:
                            col_index = columns.index('出身校') + 1
                            cell = ws.cell(row=row_index, column=col_index)
                            cell.fill = yellow_fill
                            cell.font = black_font
                            cell.alignment = center_alignment
            
            # 列幅自動調整
            for col_idx, col_name in enumerate(columns, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = 15
        
        excel_buffer.seek(0)
        return excel_buffer

class CSVCorrectionSystem:
    """CSV自動訂正システム（従来版）"""
    
    def __init__(self, jba_system, gemini_api_key=None):
        self.jba_system = jba_system
        self.validator = DataValidator(gemini_api_key)
    
    def process_csv_file(self, df, university_name, threshold=0.8, get_details=False):
        """CSVファイルを処理して訂正版を作成"""
        st.info(f"📊 CSVファイルを処理中... ({len(df)}行)")
        st.write(f"🔍 処理開始: 大学名={university_name}, 閾値={threshold}, 詳細取得={get_details}")
        
        results = []
        corrections = []
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for index, row in df.iterrows():
            progress = (index + 1) / len(df)
            progress_bar.progress(progress)
            status_text.text(f"処理中: {index + 1}/{len(df)} - {row.get('選手名', row.get('氏名', 'Unknown'))}")
            
            st.write(f"🔍 行 {index + 1} を処理中...")
            
            # 選手名のみを取得
            player_name = None
            name_columns = ['選手名', '氏名', 'name', 'Name']
            
            for col in name_columns:
                if col in df.columns and pd.notna(row[col]):
                    player_name = str(row[col]).strip()
                    st.write(f"  - 選手名取得: {player_name} (カラム: {col})")
                    break
            
            if not player_name:
                st.warning(f"  - 選手名が取得できませんでした")
                results.append({
                    'index': index,
                    'original_data': row.to_dict(),
                    'status': 'missing_data',
                    'message': '選手名が不足しています',
                    'correction': None
                })
                continue
            
            # JBAデータベースとの照合
            verification_result = self.jba_system.verify_player_info(
                player_name, None, university_name, get_details, threshold
            )
            
            result = {
                'index': index,
                'original_data': row.to_dict(),
                'verification_result': verification_result,
                'status': verification_result['status']
            }
            
            # 完全一致の場合
            if verification_result['status'] == 'match':
                if get_details and 'jba_data' in verification_result:
                    jba_data = verification_result['jba_data']
                    is_valid, validation_issues, school_corrections = self.validator.validate_player_data(jba_data)
                    
                    corrected_data = row.to_dict().copy()
                    
                    # JBA情報を追加
                    if 'height' in jba_data and jba_data['height']:
                        corrected_data['身長'] = f"{jba_data['height']}cm"
                    if 'weight' in jba_data and jba_data['weight']:
                        corrected_data['体重'] = f"{jba_data['weight']}kg"
                    if 'position' in jba_data and jba_data['position']:
                        corrected_data['ポジション'] = jba_data['position']
                    if 'school' in jba_data and jba_data['school']:
                        if 'school' in school_corrections:
                            corrected_data['出身校'] = school_corrections['school']
                            result['school_correction'] = f"{jba_data['school']} → {school_corrections['school']}"
                        else:
                            corrected_data['出身校'] = jba_data['school']
                    if 'grade' in jba_data and jba_data['grade']:
                        corrected_data['学年'] = jba_data['grade']
                    if 'uniform_number' in jba_data and jba_data['uniform_number']:
                        corrected_data['背番号'] = jba_data['uniform_number']
                    
                    if not is_valid:
                        result['validation_issues'] = validation_issues
                        result['message'] = f'JBAデータベースと完全一致（詳細情報追加）⚠️ 異常値検出: {", ".join(validation_issues)}'
                    else:
                        result['message'] = 'JBAデータベースと完全一致（詳細情報追加）'
                    
                    result['correction'] = corrected_data
                else:
                    result['correction'] = None
                    result['message'] = 'JBAデータベースと完全一致'
            
            # 部分一致の場合
            elif verification_result['status'] == 'partial_match':
                jba_data = verification_result['jba_data']
                similarity = verification_result.get('similarity', 0.0)
                
                corrected_data = row.to_dict().copy()
                
                if get_details:
                    if 'height' in jba_data and jba_data['height']:
                        corrected_data['身長'] = f"{jba_data['height']}cm"
                    if 'weight' in jba_data and jba_data['weight']:
                        corrected_data['体重'] = f"{jba_data['weight']}kg"
                    if 'position' in jba_data and jba_data['position']:
                        corrected_data['ポジション'] = jba_data['position']
                    if 'school' in jba_data and jba_data['school']:
                        corrected_data['出身校'] = jba_data['school']
                    if 'grade' in jba_data and jba_data['grade']:
                        corrected_data['学年'] = jba_data['grade']
                    if 'uniform_number' in jba_data and jba_data['uniform_number']:
                        corrected_data['背番号'] = jba_data['uniform_number']
                
                result['correction'] = corrected_data
                result['message'] = f"部分一致: {jba_data['name']} (類似度: {similarity:.3f}) - 手動確認推奨"
            
            # 一致なしの場合
            else:
                result['correction'] = None
                result['message'] = verification_result.get('message', '照合できませんでした')
            
            results.append(result)
        
        progress_bar.progress(1.0)
        status_text.text("✅ 処理完了")
        
        return results, corrections
    
    def create_corrected_csv(self, df, results):
        """訂正版CSVを作成（訂正部分を赤字で表示）"""
        corrected_df = df.copy()
        
        # 訂正を適用
        for result in results:
            if result['correction']:
                index = result['index']
                corrected_data = result['correction']
                
                # 各カラムを更新
                for col, value in corrected_data.items():
                    if col in corrected_df.columns:
                        # 元の値と異なる場合のみ訂正
                        original_value = corrected_df.at[index, col]
                        if original_value != value:
                            # 訂正された値を赤字で表示
                            corrected_df.at[index, col] = f"🔴 {value}"
        
        return corrected_df

def main():
    """メイン関数"""
    # セッション状態の初期化
    if 'jba_logged_in' not in st.session_state:
        st.session_state.jba_logged_in = False
    if 'jba_system' not in st.session_state:
        st.session_state.jba_system = None
    if 'csv_system' not in st.session_state:
        st.session_state.csv_system = None
    if 'uploaded_df' not in st.session_state:
        st.session_state.uploaded_df = None
    if 'university_name' not in st.session_state:
        st.session_state.university_name = ""
    if 'threshold' not in st.session_state:
        st.session_state.threshold = 0.8
    if 'get_details' not in st.session_state:
        st.session_state.get_details = False
    
    st.title("🏀 CSV自動訂正システム")
    st.markdown("**JBAデータベースと照合してCSVファイルを自動訂正します**")
    
    # カスタムCSS
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(90deg, #2563eb, #3b82f6);
        padding: 2rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .card {
        background: white;
        padding: 1.5rem;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        margin-bottom: 1rem;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # サイドバー
    with st.sidebar:
        st.header("🔐 JBAログイン情報")
        email = st.text_input("JBAメールアドレス", type="default")
        password = st.text_input("JBAパスワード", type="password")
        
        if st.button("JBAにログイン", type="primary"):
            if email and password:
                # JBAシステムの初期化
                if st.session_state.jba_system is None:
                    st.session_state.jba_system = JBAVerificationSystem()
                
                if st.session_state.jba_system.login(email, password):
                    st.session_state.jba_logged_in = True
                    st.success("ログイン成功")
                else:
                    st.session_state.jba_logged_in = False
                    st.error("ログイン失敗")
            else:
                st.error("ログイン情報を入力してください")
        
        # ログイン状態の表示
        if st.session_state.jba_logged_in:
            st.success("✅ JBAにログイン済み")
        else:
            st.warning("⚠️ JBAにログインしてください")
        
        # 内部設定（表示なし）
        threshold = 1.0  # 完全一致のみ
        get_details = True  # 常にオン
        # AI機能は使用しない
        use_parallel_processing = True  # 並列処理を使用
        max_workers = 5  # 並列スレッド数
    
    # システム初期化
    if 'jba_system' not in st.session_state:
        st.session_state.jba_system = JBAVerificationSystem()
    
    # ログイン状態の復元チェック
    if 'jba_logged_in' not in st.session_state:
        st.session_state.jba_logged_in = False
    
    # CSVシステムを毎回更新（APIキーの変更に対応）
    if st.session_state.jba_system is not None:
        if use_parallel_processing:
            st.session_state.csv_system = FastCSVCorrectionSystem(
                st.session_state.jba_system, 
                None,  # AI機能は使用しない
                max_workers=max_workers
            )
        else:
            st.session_state.csv_system = CSVCorrectionSystem(st.session_state.jba_system, None)  # AI機能は使用しない
    else:
        st.session_state.csv_system = None
    
    # メインコンテンツ
    st.header("📄 CSVファイル処理")
    
    # 大学名入力
    university_name = st.text_input("大学名", placeholder="例: 白鴎大学", help="JBAデータベースに登録されている大学名を入力してください")
    
    # CSVファイルアップロード
    uploaded_file = st.file_uploader(
        "CSVファイルをアップロード",
        type=['csv'],
        help="選手名が含まれるCSVファイルをアップロードしてください"
    )
    
    if uploaded_file is not None:
        try:
            # CSVファイルを読み込み（複数のエンコーディングを試行）
            encodings = ['utf-8', 'shift_jis', 'cp932', 'utf-8-sig', 'iso-2022-jp']
            df = None
            
            for encoding in encodings:
                try:
                    uploaded_file.seek(0)  # ファイルポインタをリセット
                    df = pd.read_csv(uploaded_file, encoding=encoding)
                    st.success(f"✅ CSVファイルを読み込みました ({len(df)}行) - エンコーディング: {encoding}")
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                st.error("❌ CSVファイルの文字エンコーディングが判別できませんでした。ファイルをUTF-8で保存し直してください。")
                st.stop()
            
            # 背番号関連の列を削除
            columns_to_drop = [col for col in df.columns if '背番号' in col or 'uniform' in col.lower()]
            if columns_to_drop:
                st.info(f"🔧 背番号関連の列を削除します: {columns_to_drop}")
                df = df.drop(columns=columns_to_drop)
            
            # capカラムを除外（無視）
            if 'cap' in df.columns:
                st.info("📝 「cap」カラムは無視します")
                df = df.drop(columns=['cap'])
            
            # データプレビュー
            st.subheader("📊 データプレビュー")
            st.dataframe(df.head())
            
            # カラム情報
            st.subheader("📋 カラム情報")
            st.write(f"カラム数: {len(df.columns)}")
            st.write(f"カラム名: {list(df.columns)}")
            
            # 処理実行
            if st.button("🚀 自動訂正を実行", type="primary"):
                if not st.session_state.jba_logged_in:
                    st.error("❌ 先にJBAにログインしてください")
                elif not university_name:
                    st.error("❌ 大学名を入力してください")
                elif st.session_state.csv_system is None:
                    st.error("❌ CSVシステムが初期化されていません。JBAにログインしてください。")
                else:
                    # デバッグ情報を表示
                    st.info(f"🔍 デバッグ情報:")
                    st.write(f"- 大学名: {university_name}")
                    st.write(f"- データ行数: {len(df)}")
                    st.write(f"- カラム名: {list(df.columns)}")
                    st.write(f"- JBAログイン状態: {st.session_state.jba_logged_in}")
                    st.write(f"- 類似度閾値: {threshold}")
                    st.write(f"- 詳細情報取得: {get_details}")
                    
                    # 最初の数行のデータを表示
                    st.write("**最初の3行のデータ:**")
                    st.dataframe(df.head(3))
                    
                    # CSV処理実行
                    if use_parallel_processing:
                        results = st.session_state.csv_system.process_csv_file_parallel(
                            df, university_name, threshold
                        )
                        corrections = []  # 並列処理版では corrections は別途処理
                    else:
                        results, corrections = st.session_state.csv_system.process_csv_file(
                            df, university_name, threshold, get_details
                        )
                    
                    # 結果表示
                    st.subheader("📊 処理結果")
                    
                    # 訂正ありの件数をカウント
                    total_records = len(results)
                    matched_count = sum(1 for r in results if r['status'] == 'match')
                    partial_match_count = sum(1 for r in results if r['status'] == 'partial_match')
                    has_correction_count = sum(1 for r in results if r.get('has_correction', False))
                    warnings_count = sum(len(r.get('validation_warnings', [])) for r in results)
                    not_found_count = sum(1 for r in results if r['status'] == 'not_found')
                    
                    col1, col2, col3, col4, col5, col6 = st.columns(6)
                    with col1:
                        st.metric("全件数", total_records)
                    with col2:
                        st.metric("JBA一致", matched_count)
                    with col3:
                        st.metric("部分一致", partial_match_count)
                    with col4:
                        st.metric("訂正あり", has_correction_count)  # 実際に訂正があった件数
                    with col5:
                        st.metric("⚠️ 警告", warnings_count)
                    with col6:
                        st.metric("JBA登録なし", not_found_count)
                    
                    # 訂正版CSVを作成
                    corrected_df = st.session_state.csv_system.create_corrected_csv(df, results)
                    
                    # 色付きExcelを作成
                    excel_buffer = st.session_state.csv_system.create_colored_excel(corrected_df, results)
                    
                    # ダウンロードボタン
                    st.download_button(
                        label="📊 修正版Excel（色付け）をダウンロード",
                        data=excel_buffer.getvalue(),
                        file_name=f"corrected_{uploaded_file.name.replace('.csv', '.xlsx')}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # 詳細結果表示
                    st.subheader("📋 詳細結果")
                    
                    # タブで結果を分ける
                    tab1, tab2, tab3, tab4 = st.tabs(["全詳細", "警告一覧", "JBA登録なし", "訂正あり"])
                    
                    with tab1:
                        st.write(f"**全詳細情報**")
                        for i, result in enumerate(results):
                            player_name = result['original_data'].get('選手名', 'Unknown')
                            status_emoji = {
                                'match': '✅',
                                'partial_match': '🔶',
                                'not_found': '❌',
                                'missing_data': '⚠️'
                            }.get(result['status'], '❓')
                            
                            with st.expander(f"{status_emoji} {i+1}. {player_name}"):
                                st.write(f"状態: {result['status']}")
                                if result.get('has_correction'):
                                    st.write("修正内容:")
                                    st.json(result['corrections'])
                                if result.get('validation_warnings'):
                                    st.write("警告:")
                                    for warning in result['validation_warnings']:
                                        st.warning(warning)
                    
                    with tab2:
                        warning_results = [r for r in results if r.get('validation_warnings')]
                        if warning_results:
                            st.write(f"**警告: {len(warning_results)}件**")
                            for result in warning_results:
                                player_name = result['original_data'].get('選手名', 'Unknown')
                                with st.expander(f"⚠️ {player_name}"):
                                    for warning in result['validation_warnings']:
                                        st.warning(warning)
                        else:
                            st.success("警告はありません")
                    
                    with tab3:
                        not_found_results = [r for r in results if r['status'] == 'not_found']
                        if not_found_results:
                            st.write(f"**JBA登録なし: {len(not_found_results)}件**")
                            for result in not_found_results:
                                player_name = result['original_data'].get('選手名', 'Unknown')
                                st.warning(f"❌ {player_name}")
                        else:
                            st.success("全て発見されました")
                    
                    with tab4:
                        # 訂正ありの行のみを表示
                        correction_results = [r for r in results if r.get('has_correction', False)]
                        if correction_results:
                            st.write(f"**訂正あり: {len(correction_results)}件**")
                            for result in correction_results:
                                player_name = result['original_data'].get('選手名', 'Unknown')
                                with st.expander(f"🔧 {player_name} (行{result['index']+1})"):
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.write("**修正前:**")
                                        st.json(result['original_data'])
                                    with col2:
                                        st.write("**修正内容:**")
                                        st.json(result['corrections'])
                        else:
                            st.info("訂正ありのデータはありません")
                    
        
        except Exception as e:
            st.error(f"❌ CSVファイルの読み込みに失敗しました: {str(e)}")

if __name__ == "__main__":
    main()
